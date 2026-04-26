"""
診斷：為什麼 270 位病患沒有醫師？
1. 兩份 Excel 各有多少 row 有填醫師
2. 兩份合併後，有多少 unique 病人 (姓名) 有醫師資訊
3. patients-import.json 裡有多少病人能跟 Excel match
"""
import re
from collections import defaultdict
from pathlib import Path
import json

from openpyxl import load_workbook

import os
MAIN_XLSX = os.environ.get('ALIGNER_MAIN_XLSX') or r'C:\Users\YOUR_USER\Downloads\下單紀錄.xlsx'
SUPP_XLSX = os.environ.get('ALIGNER_SUPP_XLSX') or r'C:\Users\YOUR_USER\Downloads\補充下單紀錄.xlsx'
PATIENTS_JSON = Path('dev-data/patients-import.json')


def normalize_name(raw):
    s = str(raw).strip()
    m = re.match(r'^(.+?)(\d{6,7})$', s)
    if not m:
        return s, None
    name = m.group(1).strip()
    digits = m.group(2)
    if len(digits) == 6:
        roc_y, mo, d = int(digits[:2]), int(digits[2:4]), int(digits[4:6])
    else:
        roc_y, mo, d = int(digits[:3]), int(digits[3:5]), int(digits[5:7])
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return s, None
    return name, f'{roc_y + 1911:04d}-{mo:02d}-{d:02d}'


def scan(path, label):
    wb = load_workbook(path, data_only=True)
    sh = wb.active
    print(f'\n=== {label} ({path}) ===')
    print(f'總 row 數: {sh.max_row}')

    rows_with_name = 0
    rows_with_doctor = 0
    rows_doctor_missing = 0
    distinct_doctors = defaultdict(int)
    name_to_doctor = {}  # clean_name → set of doctors seen
    name_doctor_track = defaultdict(set)

    for r in range(2, sh.max_row + 1):
        name = sh.cell(row=r, column=2).value
        doctor = sh.cell(row=r, column=3).value
        if name:
            rows_with_name += 1
            clean_name, _ = normalize_name(name)
            if doctor and str(doctor).strip():
                rows_with_doctor += 1
                d = str(doctor).strip()
                distinct_doctors[d] += 1
                name_doctor_track[clean_name].add(d)
            else:
                rows_doctor_missing += 1

    print(f'有姓名的 row: {rows_with_name}')
    print(f'  其中有填醫師: {rows_with_doctor} ({rows_with_doctor/rows_with_name*100:.1f}%)')
    print(f'  其中無醫師:    {rows_doctor_missing}')
    print(f'\n出現過的醫師 (count):')
    for d, c in sorted(distinct_doctors.items(), key=lambda x: -x[1]):
        print(f'  {d}: {c}')
    print(f'\n獨立病人姓名數: {len(name_doctor_track)}')
    print(f'  其中至少 1 row 有醫師: {sum(1 for v in name_doctor_track.values() if v)}')
    return name_doctor_track


def main():
    main_map = scan(MAIN_XLSX, '主下單紀錄')
    supp_map = scan(SUPP_XLSX, '補充下單紀錄')

    # 合併
    print('\n=== 合併兩份 Excel ===')
    merged = defaultdict(set)
    for n, ds in main_map.items():
        merged[n].update(ds)
    for n, ds in supp_map.items():
        merged[n].update(ds)

    have_dr = sum(1 for v in merged.values() if v)
    print(f'合併後獨立病人數: {len(merged)}')
    print(f'  其中至少有 1 個醫師資訊: {have_dr}')
    print(f'  完全沒醫師資訊:           {len(merged) - have_dr}')

    # 跟 patients-import.json 比對
    pdata = json.loads(PATIENTS_JSON.read_text(encoding='utf-8'))
    pat_names = set(p['name'] for p in pdata['patients'])
    excel_names = set(merged.keys())

    in_both = pat_names & excel_names
    pat_only = pat_names - excel_names
    excel_only = excel_names - pat_names

    print(f'\n=== Patient ↔ Excel 名字交集 ===')
    print(f'patients-import.json 有 {len(pat_names)} 位')
    print(f'兩份 Excel 共有        {len(excel_names)} 位獨立姓名')
    print(f'兩邊都有 (可 match):    {len(in_both)}')
    print(f'只在 patients (Excel 沒登): {len(pat_only)}')
    print(f'只在 Excel (folder 沒掃到): {len(excel_only)}')

    # 兩邊都有但 Excel 沒醫師的人
    in_both_have_dr = sum(1 for n in in_both if merged[n])
    print(f'\n兩邊都有 + Excel 有醫師: {in_both_have_dr}')
    print(f'兩邊都有 + Excel 沒醫師: {len(in_both) - in_both_have_dr}')

    # 印出 Excel 沒登的 patient 列表（前 20 個）
    if pat_only:
        print(f'\n=== Excel 沒登的病患 (前 30 個) ===')
        for n in sorted(pat_only)[:30]:
            print(f'  {n[0]+"*"*(len(n)-1)}')


if __name__ == '__main__':
    main()
