"""掃 Excel 全部 sheet、找 B 欄（姓名）疑似汙染的 row。

判定規則（任一命中即 flag）：
1. 含 "下單" / "收到" / "寄出" — 是 batch 備註關鍵字
2. 含 aligner range pattern (UL/U/L 開頭數字-數字)
3. 含 step / IPR / ATT / BT — batch 治療備註
4. 長度 > 8 字 — 真人姓名通常 2-4 字、5-6 很罕見
5. 含括號 ( ) — 名字不會有
6. 數字過多 (除尾巴民國生日 6-7 位數)
"""
import sys
import io
import re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

XLSX = r'D:\診所nas 矯正追蹤\SynologyDrive\下單Excel\2026_01_28~診所接手生產列印名單.xlsx'

# 觸發 flag 的關鍵字 / pattern
SUSPECT_KEYWORDS = ['下單', '收到', '寄出', 'step', 'IPR', 'ATT', '已領', '無指示', 'END', '結束']
ALIGNER_PATTERN = re.compile(r'\b[UL]+\d+(?:-\d+)?')
EXCESSIVE_BRACKETS = re.compile(r'[()（）]')

def is_suspect(b_value):
    if b_value is None:
        return None
    s = str(b_value).strip()
    if not s:
        return None
    reasons = []
    # 末尾民國生日 strip 掉再判長度
    s_no_birth = re.sub(r'\d{6,7}$', '', s).strip()
    # 1. 太長
    if len(s_no_birth) > 8:
        reasons.append(f'長 ({len(s_no_birth)} 字)')
    # 2. 含 batch 關鍵字
    hit_keywords = [k for k in SUSPECT_KEYWORDS if k in s]
    if hit_keywords:
        reasons.append(f'含關鍵字 {hit_keywords}')
    # 3. aligner pattern
    if ALIGNER_PATTERN.search(s):
        reasons.append('含 aligner range')
    # 4. 括號
    if EXCESSIVE_BRACKETS.search(s):
        reasons.append('含括號')
    return reasons if reasons else None

wb = load_workbook(XLSX, data_only=True)

# 只關心 「牙套下單」、「生產資料庫」 兩個主要 sheet
for sh_name in ['牙套下單', '生產資料庫']:
    if sh_name not in wb.sheetnames:
        continue
    sh = wb[sh_name]
    print(f'═══════════ sheet: {sh_name} (掃 {sh.max_row} rows) ═══════════')
    flagged = []
    for r in range(2, sh.max_row + 1):
        b = sh.cell(row=r, column=2).value
        reasons = is_suspect(b)
        if reasons:
            flagged.append((r, str(b)[:60], reasons))
    print(f'  發現 {len(flagged)} 筆疑似異常\n')
    for r, b_short, reasons in flagged:
        print(f'  row {r:4d} | {b_short}')
        print(f'           reason: {", ".join(reasons)}')
        # 順便看 C / H 欄、判斷真實姓名線索
        c = sh.cell(row=r, column=3).value
        h = sh.cell(row=r, column=8).value
        print(f'           C(醫師): {c}    H(aligner): {str(h)[:50] if h else "—"}')
        # 看下一筆有沒有真實姓名（同 patient group 第二行通常 name 對）
        next_b = sh.cell(row=r+1, column=2).value
        if next_b:
            print(f'           ↓ next row B: {next_b}  ← 可能是真實姓名')
        print()
    print()
