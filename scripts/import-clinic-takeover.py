# Clinic-Takeover Excel 匯入 → JSON
#
# 取代原本的 import-excel-orders.py + import-supplementary-orders.py 雙腳本。
# 從「2026_01_28~診所接手生產列印名單.xlsx」這類**單一檔案兩個分頁**讀資料：
#   - Sheet「生產資料庫」(1 row / patient) — 病患總表，最多 1 筆 order
#   - Sheet「牙套下單」(多 row / patient) — 下過 2+ 次單者的所有 orders
#
# 路徑解析（環境變數 → 自動偵測 → 名稱比對）：
#   1. ALIGNER_TAKEOVER_XLSX env 指定檔
#   2. ${ALIGNER_EXCEL_FOLDER} 內找含「生產資料庫」+「牙套下單」兩個 sheet 的 .xlsx
#   3. 自動偵測 D:\矯正\下單Excel\ 或 C:\矯正\下單Excel\
#
# 產出：dev-data/excel-orders.json + dev-data/excel-patient-updates.json
# (跟舊的 import-excel-orders.py 同 schema、可被 src/seed.ts 跟 lib/reapply-excel.ts 直接吃)
#
# 執行：python scripts/import-clinic-takeover.py

import json
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PATIENTS_JSON = Path('dev-data/patients-import.json')
OUT_ORDERS = Path('dev-data/excel-orders.json')
OUT_PATIENT_UPDATES = Path('dev-data/excel-patient-updates.json')
# v0.4.14 加：第 4 個分頁（轉隱適美 等轉品牌資料）的 dump、給 DoctorBackfillSection 階段 3 fallback 用
OUT_TRANSFERRED = Path('dev-data/excel-transferred.json')
TRANSFER_SHEET_INDEX = 3  # 0-based: 第 4 個分頁

LAB_MAP = {
    'invisalign': '隱適美',
    'riyue': '美鉑',
    'zenyum': '美鉑',
    'retainer': '美鉑',
}

PROGRESS_VALID = {'尚未開始', '設計中', '已下單牙套', '診所已收到牙套', '已完成'}

SHEET_MASTER = '生產資料庫'  # 1 row / patient
SHEET_ORDERS = '牙套下單'  # 多 row / patient


# ─── 路徑解析 ────────────────────────────────────────
def _resolve_xlsx():
    if os.environ.get('ALIGNER_TAKEOVER_XLSX'):
        return os.environ['ALIGNER_TAKEOVER_XLSX']

    folder = os.environ.get('ALIGNER_EXCEL_FOLDER')
    if not folder:
        for d in (r'D:\矯正\下單Excel', r'C:\矯正\下單Excel'):
            if os.path.isdir(d):
                folder = d
                break
    if not folder:
        return None

    # 找含「生產資料庫」+「牙套下單」兩個 sheet 的 .xlsx
    for f in os.listdir(folder):
        if not f.endswith('.xlsx'):
            continue
        full = os.path.join(folder, f)
        try:
            wb = load_workbook(full, read_only=True)
            if SHEET_MASTER in wb.sheetnames and SHEET_ORDERS in wb.sheetnames:
                wb.close()
                return full
            wb.close()
        except Exception:
            continue
    return None


# ─── Helpers (跟 import-excel-orders.py 一致) ──────
def normalize_progress(v):
    if not v:
        return '尚未開始'
    s = str(v).strip()
    return s if s in PROGRESS_VALID else '尚未開始'


def normalize_consent(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('有',):
        return True
    if s in ('X', 'x', '無'):
        return False
    return None


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


def parse_aligner_range(text):
    """從 'UL12-13 備註' 抽出 'UL12-13' + '備註'。
    不是合法 range 開頭時回 ('', s) — 全文當備註、range 為空（這 row 不是真 order）"""
    if not text:
        return '', ''
    s = str(text).strip()
    m = re.match(r'^([UL]+\d+(?:-\d+)?(?:\+[UL]+\d+(?:-\d+)?)*)', s)
    if not m:
        return '', s  # ← 之前 bug：return (s, '')，把整段中文當 range
    rng = m.group(1)
    return rng, s[len(rng):].strip()


def extract_actual_order_date(text, fallback_year):
    if not text:
        return None
    m = re.search(r'(\d{1,2})/(\d{1,2})\s*下單', str(text))
    if not m:
        return None
    mo, d = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return f'{fallback_year}-{mo:02d}-{d:02d}'


def extract_design_submit_date(text, fallback_year):
    """從 'M/D送出設計檔' 抓日期。例：'4/7送出設計檔設計' → 2026-04-07"""
    if not text:
        return None
    m = re.search(r'(\d{1,2})/(\d{1,2})\s*送出設計檔', str(text))
    if not m:
        return None
    mo, d = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return f'{fallback_year}-{mo:02d}-{d:02d}'


def normalize_excel_name(raw_name):
    """處理 '王小明810413' (姓名+民國日)，回 (clean_name, birthday_iso_or_none)"""
    s = str(raw_name).strip()
    m = re.match(r'^(.+?)(\d{6,7})$', s)
    if not m:
        return s, None
    name = m.group(1).strip()
    digits = m.group(2)
    if len(digits) == 6:
        roc_y, mo, d = int(digits[:2]), int(digits[2:4]), int(digits[4:6])
    else:
        roc_y, mo, d = int(digits[:3]), int(digits[3:5]), int(digits[5:7])
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return s, None
    return name, f'{roc_y + 1911:04d}-{mo:02d}-{d:02d}'


def parse_upper_lower_totals(v):
    if v is None:
        return None, None
    m = re.match(r'^(\d+)\s*/\s*(\d+)$', str(v).strip())
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


# ─── Row 抽 dict ──────────────────────────────────────
def read_row(sh, r):
    name = sh.cell(row=r, column=2).value
    if not name:
        return None
    return {
        'r': r,
        'date': parse_date(sh.cell(row=r, column=1).value),
        'name': str(name).strip(),
        'doctor': sh.cell(row=r, column=3).value,
        'scan': sh.cell(row=r, column=4).value,
        'consent': sh.cell(row=r, column=5).value,
        'totalProgress': sh.cell(row=r, column=6).value,
        'batchType': sh.cell(row=r, column=7).value,
        'alignerRangeRaw': sh.cell(row=r, column=8).value,
        'progress': sh.cell(row=r, column=9).value,
        'notes': sh.cell(row=r, column=10).value,
        'expectedDate': sh.cell(row=r, column=15).value,
        'actualDate': sh.cell(row=r, column=17).value,
        'nextStep': sh.cell(row=r, column=18).value,
    }


def main():
    if not Path(PATIENTS_JSON).exists():
        print(f'❌ 找不到 {PATIENTS_JSON}', file=sys.stderr)
        sys.exit(1)
    excel_path = _resolve_xlsx()
    if not excel_path:
        print(
            '❌ 找不到含「生產資料庫」+「牙套下單」兩個分頁的 Excel 檔。\n'
            '   請放到 D:\\矯正\\下單Excel\\ 或 C:\\矯正\\下單Excel\\\n'
            '   或設定環境變數 ALIGNER_TAKEOVER_XLSX 指定路徑。',
            file=sys.stderr,
        )
        sys.exit(1)
    print(f'[takeover] 讀檔：{excel_path}')

    with open(PATIENTS_JSON, 'r', encoding='utf-8') as f:
        patients_data = json.load(f)

    name_to_patients = {}
    for p in patients_data['patients']:
        name_to_patients.setdefault(p['name'], []).append(p)

    wb = load_workbook(excel_path, data_only=True)

    orders = []
    patient_updates = {}  # patientId → fields
    new_patients = []
    multi_match = []
    now_iso = datetime.now().isoformat()
    max_chart = max(int(p['chartNo']) for p in patients_data['patients'])

    def next_chart():
        nonlocal max_chart
        max_chart += 1
        return f'{max_chart:04d}'

    # ── 把 group rows 處理成 patient + orders ──
    def process_group(grp):
        head = grp[0]
        raw_name = head['name']
        clean_name, hint_birthday = normalize_excel_name(raw_name)
        candidates = name_to_patients.get(clean_name, [])

        if len(candidates) == 0:
            consent = normalize_consent(head['consent'])
            upper_t, lower_t = parse_upper_lower_totals(head['totalProgress'])
            new_pat = {
                'id': str(uuid.uuid4()),
                'chartNo': next_chart(),
                'name': clean_name,
                'birthday': hint_birthday,
                'productLine': 'riyue',
                'status': 'active',
                'track': None,
                'refinementLevel': 0,
                'orderDate': head['date'],
                'startDate': None,
                'totalAlignersUpper': upper_t,
                'currentAlignerUpper': None,
                'totalAlignersLower': lower_t,
                'currentAlignerLower': None,
                'cycleDays': 14,
                'lastVisit': None,
                'nextVisit': None,
                'hasConsent': consent if consent is not None else False,
                'consentPdfPath': None,
                'scanInfo': str(head['scan']).strip()
                if head['scan'] and str(head['scan']).strip() not in ('X', 'x', '無')
                else None,
                'doctor': str(head['doctor']).strip() if head['doctor'] else None,
                'flags': [],
                'notes': '(從 Excel 匯入新增，原始資料夾不在 scan 範圍)',
                'sourceFolder': '',
                'createdAt': now_iso,
                'updatedAt': now_iso,
            }
            new_patients.append(new_pat)
            patient = new_pat
            name_to_patients[clean_name] = [new_pat]
        else:
            if len(candidates) > 1 and hint_birthday:
                matched = [c for c in candidates if c['birthday'] == hint_birthday]
                if len(matched) == 1:
                    candidates = matched
            if len(candidates) > 1:
                multi_match.append({
                    'name': raw_name,
                    'rows': [r['r'] for r in grp],
                })
            patient = candidates[0]

        # Patient 補欄更新
        upd = {}
        if head['scan'] and str(head['scan']).strip() not in ('X', 'x', '無'):
            upd['scanInfo'] = str(head['scan']).strip()
        if head['doctor']:
            upd['doctor'] = str(head['doctor']).strip()
        consent_bool = normalize_consent(head['consent'])
        if consent_bool is not None:
            upd['hasConsent'] = consent_bool
        upper_t, lower_t = parse_upper_lower_totals(head['totalProgress'])
        if upper_t is not None:
            upd['totalAlignersUpper'] = upper_t
        if lower_t is not None:
            upd['totalAlignersLower'] = lower_t
        if upd:
            patient_updates.setdefault(patient['id'], {}).update(upd)

        # 每個 row 轉成 1 筆 order
        case_year = 2026
        if head['date']:
            case_year = int(head['date'].split('-')[0])
        for row in grp:
            rng, rest = parse_aligner_range(row['alignerRangeRaw'])
            raw_text = str(row['alignerRangeRaw'] or '')
            has_design_submit = '送出設計檔' in raw_text
            if not rng and not has_design_submit:
                # 沒 UL/L 副數區間、也沒「送出設計檔」 → 跳過（純諮詢 / 退費）
                continue
            combined = f"{raw_text} {row['notes'] or ''}"
            # 「送出設計檔」row → 進度改 設計中、日期從 'M/D送出設計檔' 抽
            if has_design_submit and not rng:
                actual = extract_design_submit_date(raw_text, case_year)
                progress_value = '設計中'
            else:
                actual = extract_actual_order_date(combined, case_year)
                progress_value = normalize_progress(row['progress'])
            order_date = actual or row['date'] or head['date'] or ''
            orders.append({
                'id': f"excel-{row['r']:04d}-{patient['id']}",
                'patientId': patient['id'],
                'patientChartNo': patient['chartNo'],
                'patientName': patient['name'],
                'date': order_date,
                'doctor': str(row['doctor']).strip()
                if row['doctor']
                else (str(head['doctor']).strip() if head['doctor'] else ''),
                'batchType': str(row['batchType']).strip()
                if row['batchType'] and str(row['batchType']).strip() not in ('X', 'x')
                else '',
                'alignerRange': rng,
                'progress': progress_value,
                'expectedDate': parse_date(row['expectedDate']),
                'actualDate': parse_date(row['actualDate']),
                'nextStep': str(row['nextStep']).strip() if row['nextStep'] else '',
                'notes': ' '.join(filter(None, [
                    rest,
                    str(row['notes']).strip() if row['notes'] else '',
                ])).strip(),
                'lab': LAB_MAP.get(patient['productLine'], '美鉑'),
                'createdAt': now_iso,
                'updatedAt': now_iso,
            })

    # ─── Sheet 1: 生產資料庫 (1 row / patient) ──────
    sh1 = wb[SHEET_MASTER]
    n1_processed = 0
    for r in range(2, sh1.max_row + 1):
        row = read_row(sh1, r)
        if row:
            process_group([row])
            n1_processed += 1
    print(f'[takeover] {SHEET_MASTER}: 處理 {n1_processed} 個 row')

    # ─── Sheet 2: 牙套下單 (多 row / patient, 空白列分隔) ──────
    sh2 = wb[SHEET_ORDERS]
    groups2 = []
    current = []
    for r in range(2, sh2.max_row + 1):
        row = read_row(sh2, r)
        if not row:
            if current:
                groups2.append(current)
                current = []
            continue
        current.append(row)
    if current:
        groups2.append(current)
    for grp in groups2:
        process_group(grp)
    print(f'[takeover] {SHEET_ORDERS}: 處理 {len(groups2)} 群、共 {sum(len(g) for g in groups2)} rows')

    # ─── 輸出 ─────────────────────────────────
    pat_lookup = {p['id']: p for p in patients_data['patients']}

    def enrich_update(pid, fields):
        ref = pat_lookup.get(pid, {})
        return {
            'id': pid,
            'refName': ref.get('name'),
            'refBirthday': ref.get('birthday'),
            **fields,
        }

    OUT_ORDERS.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_ORDERS, 'w', encoding='utf-8') as f:
        json.dump({'count': len(orders), 'orders': orders}, f, ensure_ascii=False, indent=2)
    with open(OUT_PATIENT_UPDATES, 'w', encoding='utf-8') as f:
        json.dump(
            {
                'count': len(patient_updates),
                'newPatientsCount': len(new_patients),
                'updates': [enrich_update(pid, fields) for pid, fields in patient_updates.items()],
                'newPatients': new_patients,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    # ─── Sheet 4 (v0.4.14): 轉品牌分頁 (轉隱適美、轉綻雅 等) ─────
    # generic header scan、不依賴固定欄位 layout
    # 抽出 patientName / doctor / birthday / sourceSheet 給 DoctorBackfillSection 用
    transferred_rows = []
    sheet_names = wb.sheetnames
    if len(sheet_names) > TRANSFER_SHEET_INDEX:
        transfer_sheet_name = sheet_names[TRANSFER_SHEET_INDEX]
        sh4 = wb[transfer_sheet_name]
        print(f'\n[takeover] 第 {TRANSFER_SHEET_INDEX + 1} 個分頁: 「{transfer_sheet_name}」 ({sh4.max_row} rows)')

        # Scan header (row 1)
        headers = {}
        for c in range(1, sh4.max_column + 1):
            v = sh4.cell(row=1, column=c).value
            if v:
                headers[c] = str(v).strip()
        print(f'[takeover]   headers: {headers}')

        NAME_KEYS = ['姓名', '病患姓名', '病人姓名', '病患']
        DOCTOR_KEYS = ['醫師', '主治醫師', '主治']
        BIRTHDAY_KEYS = ['生日', '出生', '出生年月日']

        def find_col(keys):
            for col, h in headers.items():
                for k in keys:
                    if k in h:
                        return col
            return None

        name_col = find_col(NAME_KEYS)
        doctor_col = find_col(DOCTOR_KEYS)
        birthday_col = find_col(BIRTHDAY_KEYS)

        if name_col is None:
            print(f'[takeover]   ⚠ 找不到「姓名」欄、跳過此分頁（header 沒含「姓名」/「病患」等關鍵字）')
        else:
            print(f'[takeover]   姓名 col={name_col}, 醫師 col={doctor_col}, 生日 col={birthday_col}')
            for r in range(2, sh4.max_row + 1):
                name_v = sh4.cell(row=r, column=name_col).value
                if not name_v:
                    continue
                row_data = {
                    'patientName': str(name_v).strip(),
                    'sourceSheet': transfer_sheet_name,
                    'doctor': '',
                }
                if doctor_col:
                    d = sh4.cell(row=r, column=doctor_col).value
                    if d:
                        row_data['doctor'] = str(d).strip()
                if birthday_col:
                    bday = sh4.cell(row=r, column=birthday_col).value
                    if isinstance(bday, datetime):
                        row_data['birthday'] = bday.strftime('%Y-%m-%d')
                    elif bday:
                        row_data['birthday'] = str(bday)
                transferred_rows.append(row_data)
            print(f'[takeover]   抽出 {len(transferred_rows)} rows')
    else:
        print(f'\n[takeover] Excel 不足 {TRANSFER_SHEET_INDEX + 1} 個分頁、跳過轉品牌處理 (現有 {len(sheet_names)} 個)')

    with open(OUT_TRANSFERRED, 'w', encoding='utf-8') as f:
        json.dump({'count': len(transferred_rows), 'rows': transferred_rows}, f, ensure_ascii=False, indent=2)

    # ─── 統計 ────────────────────────────────
    print()
    print('=== Takeover Import 統計 ===')
    print(f'  orders 總數: {len(orders)}')
    print(f'  patient updates: {len(patient_updates)}')
    print(f'  自動補建新病患: {len(new_patients)}')
    print(f'  同名多人警告: {len(multi_match)}')
    print(f'  轉品牌分頁 rows: {len(transferred_rows)}')
    by_progress = {}
    for o in orders:
        by_progress[o['progress']] = by_progress.get(o['progress'], 0) + 1
    print()
    print('=== Orders by progress ===')
    for k, v in sorted(by_progress.items(), key=lambda x: -x[1]):
        print(f'  {k}: {v}')


if __name__ == '__main__':
    main()
