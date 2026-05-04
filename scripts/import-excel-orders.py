# Excel 下單記錄 → JSON
#
# 讀 下單記錄.xlsx (路徑見 EXCEL_PATH 常數 / ALIGNER_MAIN_XLSX env)，產出兩份 JSON：
#   1. dev-data/excel-orders.json — Order 紀錄陣列
#   2. dev-data/excel-patient-updates.json — Patient 欄位更新（scanInfo / doctor / 總副數）
#
# 同時印出 warnings：找不到 patient match 的 Excel 行、解析異常等。
#
# 執行：python scripts/import-excel-orders.py

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import os
# 主下單紀錄 Excel 路徑：
#   1. 環境變數 ALIGNER_MAIN_XLSX 直接指定檔案
#   2. 否則找 ${EXCEL_FOLDER}\下單紀錄.xlsx (預設 D:\矯正\下單Excel\)
#   3. 最後 fallback 舊位置 C:\Users\YOUR_USER\Downloads\
def _resolve_main_xlsx():
    if os.environ.get('ALIGNER_MAIN_XLSX'):
        return os.environ['ALIGNER_MAIN_XLSX']
    folder = os.environ.get('ALIGNER_EXCEL_FOLDER')
    if not folder:
        # 自動偵測 D 或 C 槽
        for d in (r'D:\矯正\下單Excel', r'C:\矯正\下單Excel'):
            if os.path.isdir(d):
                folder = d
                break
    if folder:
        candidate = os.path.join(folder, '下單紀錄.xlsx')
        if os.path.isfile(candidate):
            return candidate
    return r'C:\Users\YOUR_USER\Downloads\下單紀錄.xlsx'

EXCEL_PATH = _resolve_main_xlsx()
PATIENTS_JSON = Path('dev-data/patients-import.json')
OUT_ORDERS = Path('dev-data/excel-orders.json')
OUT_PATIENT_UPDATES = Path('dev-data/excel-patient-updates.json')

LAB_MAP = {
    'invisalign': '隱適美',
    'riyue': '美鉑',
    'zenyum': '美鉑',
    'retainer': '美鉑',
}

PROGRESS_VALID = {'尚未開始', '已下單牙套', '診所已收到牙套', '已完成'}


def normalize_progress(v):
    if not v:
        return '尚未開始'
    s = str(v).strip()
    return s if s in PROGRESS_VALID else '尚未開始'


def normalize_consent(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('有',):
        return True
    if s in ('X', 'x'):
        return False
    return None  # 其他怪值 (e.g. "陳朋友") 不動，讓現有 patient.hasConsent 保留


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


def parse_aligner_range(text):
    """從 'UL12-13 2/23下單 ...' 抽出 'UL12-13' 與剩下備註。"""
    if not text:
        return '', ''
    s = str(text).strip()
    # 匹配開頭的 [UL]+ 數字-數字 (可選+UL數字-數字)
    m = re.match(r'^([UL]+\d+(?:-\d+)?(?:\+[UL]+\d+(?:-\d+)?)*)', s)
    if not m:
        return s, ''
    rng = m.group(1)
    rest = s[len(rng):].strip()
    return rng, rest


def extract_actual_order_date(text, fallback_year):
    """從備註裡抓 'M/D下單' 或 '(M/D下單)'，回傳 ISO 日期字串。
    例：'UL14-17 2/23下單' → 2026-02-23
        'UL18-23 (4/22下單)' → 2026-04-22
        '無下單字樣' → None
    """
    if not text:
        return None
    s = str(text)
    m = re.search(r'(\d{1,2})/(\d{1,2})\s*下單', s)
    if not m:
        return None
    mo = int(m.group(1))
    d = int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return f'{fallback_year}-{mo:02d}-{d:02d}'


def normalize_excel_name(raw_name):
    """處理 '王小明810413' 這種 (姓名 + 民國生日) 的格式。
    回傳 (clean_name, birthday_iso_or_none)。
    """
    s = str(raw_name).strip()
    # 末尾連續 6-7 位數字
    m = re.match(r'^(.+?)(\d{6,7})$', s)
    if not m:
        return s, None
    name = m.group(1).strip()
    digits = m.group(2)
    # 解民國 → 西元
    if len(digits) == 6:
        roc_y = int(digits[:2])
        mo = int(digits[2:4])
        d = int(digits[4:6])
    else:
        roc_y = int(digits[:3])
        mo = int(digits[3:5])
        d = int(digits[5:7])
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return s, None
    year = roc_y + 1911
    return name, f'{year:04d}-{mo:02d}-{d:02d}'


def parse_upper_lower_totals(v):
    """Excel 的「總副數」欄寫成 'X/Y'，語意是「上顎副數 / 下顎副數」。
    e.g. '14/7' → (upper=14, lower=7)
    回傳 (upper, lower) 或 (None, None) 若無法解析。
    """
    if v is None:
        return None, None
    s = str(v).strip()
    m = re.match(r'^(\d+)\s*/\s*(\d+)$', s)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def main():
    if not Path(PATIENTS_JSON).exists():
        print(f'❌ 找不到 {PATIENTS_JSON}', file=sys.stderr)
        sys.exit(1)
    with open(PATIENTS_JSON, 'r', encoding='utf-8') as f:
        patients_data = json.load(f)

    # name → patient (含同名警告)
    name_to_patients = {}
    for p in patients_data['patients']:
        name_to_patients.setdefault(p['name'], []).append(p)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    sh = wb.active

    # Group consecutive rows by patient name (空白列 = 分隔)
    groups = []
    current = []
    for r in range(2, sh.max_row + 1):
        name = sh.cell(row=r, column=2).value
        if not name:
            if current:
                groups.append(current)
                current = []
            continue
        row = {
            'r': r,
            'date': parse_date(sh.cell(row=r, column=1).value),
            'name': str(name).strip(),
            'doctor': sh.cell(row=r, column=3).value,
            'scan': sh.cell(row=r, column=4).value,
            'consent': sh.cell(row=r, column=5).value,
            'totalProgress': sh.cell(row=r, column=6).value,
            'batchType': sh.cell(row=r, column=7).value,
            'alignerRangeRaw': sh.cell(row=r, column=8).value,
            'progress': sh.cell(row=r, column=9).value,
            'notes': sh.cell(row=r, column=10).value,
            'expectedDate': sh.cell(row=r, column=15).value,
            'actualDate': sh.cell(row=r, column=17).value,
            'nextStep': sh.cell(row=r, column=18).value,
        }
        current.append(row)
    if current:
        groups.append(current)

    orders = []
    patient_updates = {}  # patientId → fields to update
    new_patients = []  # 自動補建的新病患
    multi_match_warnings = []
    parse_warnings = []

    now_iso = datetime.now().isoformat()
    import uuid

    # 計算下一個可用 chartNo (from existing patients)
    max_chart = max(int(p['chartNo']) for p in patients_data['patients'])

    def next_chart():
        nonlocal max_chart
        max_chart += 1
        return f'{max_chart:04d}'

    for grp in groups:
        head = grp[0]
        raw_name = head['name']
        clean_name, hint_birthday = normalize_excel_name(raw_name)

        # 嘗試 match patient
        candidates = name_to_patients.get(clean_name, [])
        if len(candidates) == 0:
            # 自動建一個新 patient
            consent = normalize_consent(head['consent'])
            upper_t, lower_t = parse_upper_lower_totals(head['totalProgress'])
            new_pat = {
                'id': str(uuid.uuid4()),
                'chartNo': next_chart(),
                'name': clean_name,
                'birthday': hint_birthday,
                'productLine': 'riyue',  # 預設日月辰心
                'status': 'active',
                'track': None,  # 由 seed.ts 從 orders 推
                'refinementLevel': 0,  # 同上
                'orderDate': head['date'],
                'startDate': None,
                'totalAlignersUpper': upper_t,
                'currentAlignerUpper': None,  # current 從 orders 推算，不從 Excel col 6 讀
                'totalAlignersLower': lower_t,
                'currentAlignerLower': None,
                'cycleDays': 14,
                'lastVisit': None,
                'nextVisit': None,
                'hasConsent': consent if consent is not None else False,
                'consentPdfPath': None,
                'scanInfo': str(head['scan']).strip() if head['scan'] and str(head['scan']).strip() not in ('X', 'x') else None,
                'doctor': str(head['doctor']).strip() if head['doctor'] else None,
                'flags': [],
                'notes': '(從 Excel 匯入新增，原始資料夾不在 scan 範圍)',
                'sourceFolder': '',
                'createdAt': now_iso,
                'updatedAt': now_iso,
            }
            new_patients.append(new_pat)
            patient = new_pat
            name_to_patients[clean_name] = [new_pat]
        else:
            if len(candidates) > 1:
                # 用 birthday hint 二次篩選
                if hint_birthday:
                    matched = [c for c in candidates if c['birthday'] == hint_birthday]
                    if len(matched) == 1:
                        candidates = matched
                if len(candidates) > 1:
                    multi_match_warnings.append({
                        'name': raw_name,
                        'rows': [r['r'] for r in grp],
                        'candidates': [{'chartNo': c['chartNo'], 'birthday': c['birthday']} for c in candidates],
                    })
            patient = candidates[0]

        # ── Patient 層級更新 (取 head 那行的值) ──
        updates = {}
        if head['scan'] and str(head['scan']).strip() not in ('X', 'x'):
            updates['scanInfo'] = str(head['scan']).strip()
        if head['doctor']:
            updates['doctor'] = str(head['doctor']).strip()
        consent_bool = normalize_consent(head['consent'])
        if consent_bool is not None:
            updates['hasConsent'] = consent_bool
        upper_t, lower_t = parse_upper_lower_totals(head['totalProgress'])
        if upper_t is not None:
            updates['totalAlignersUpper'] = upper_t
        if lower_t is not None:
            updates['totalAlignersLower'] = lower_t
        # current 從 orders 推算 (在 reapply-excel.ts 做)，不從 Excel col 6 讀
        if updates:
            patient_updates.setdefault(patient['id'], {}).update(updates)

        # ── 每行轉成一筆 Order ──
        # 案件年份（用 head 的 date 推定，預設 2026）
        case_year = 2026
        if head['date']:
            case_year = int(head['date'].split('-')[0])
        for row in grp:
            rng, rest = parse_aligner_range(row['alignerRangeRaw'])
            if not rng and not rest:
                continue  # 完全空的列略過
            # 抓「M/D下單」當實際下單日，沒抓到 fallback 用 Excel 第一欄
            combined = f"{rest or ''} {row['notes'] or ''}"
            actual_order_date = extract_actual_order_date(combined, case_year)
            order_date = actual_order_date or row['date'] or head['date'] or ''
            orders.append({
                'id': f"excel-{row['r']:04d}-{patient['id']}",  # deterministic id
                'patientId': patient['id'],
                'patientChartNo': patient['chartNo'],
                'patientName': patient['name'],
                'date': order_date,
                'doctor': str(row['doctor']).strip() if row['doctor'] else (str(head['doctor']).strip() if head['doctor'] else ''),
                'batchType': str(row['batchType']).strip() if row['batchType'] and str(row['batchType']).strip() != 'X' else '',
                'alignerRange': rng,
                'progress': normalize_progress(row['progress']),
                'expectedDate': parse_date(row['expectedDate']),
                'actualDate': parse_date(row['actualDate']),
                'nextStep': str(row['nextStep']).strip() if row['nextStep'] else '',
                'notes': ' '.join(filter(None, [
                    rest,
                    str(row['notes']).strip() if row['notes'] else '',
                ])).strip(),
                'lab': LAB_MAP.get(patient['productLine'], '美鉑'),
                'createdAt': now_iso,
                'updatedAt': now_iso,
            })

    # 寫出
    # update entry 額外帶上 name + birthday，以便 reapply 時若 id 對不上 (scan 重產 UUID) 可 fallback
    pat_lookup = {p['id']: p for p in patients_data['patients']}

    def enrich_update(pid, fields):
        ref = pat_lookup.get(pid, {})
        return {
            'id': pid,
            'refName': ref.get('name'),
            'refBirthday': ref.get('birthday'),
            **fields,
        }

    OUT_ORDERS.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_ORDERS, 'w', encoding='utf-8') as f:
        json.dump({'count': len(orders), 'orders': orders}, f, ensure_ascii=False, indent=2)
    with open(OUT_PATIENT_UPDATES, 'w', encoding='utf-8') as f:
        json.dump(
            {
                'count': len(patient_updates),
                'newPatientsCount': len(new_patients),
                'updates': [enrich_update(pid, fields) for pid, fields in patient_updates.items()],
                'newPatients': new_patients,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    # 統計報告
    print('=== Import 統計 ===')
    print(f'  Excel 群組數 (病患): {len(groups)}')
    print(f'  匹配成功，產出 orders: {len(orders)}')
    print(f'  patient updates: {len(patient_updates)}')
    print(f'  自動補建新病患: {len(new_patients)}')
    print(f'  同名多人警告: {len(multi_match_warnings)}')
    print()
    if new_patients:
        print('=== 自動補建的新病患 ===')
        for n in new_patients:
            print(f"  {n['chartNo']} {n['name']} 生日={n['birthday']} 醫師={n['doctor']}")
        print()
    if multi_match_warnings:
        print('=== 同名警告 ===')
        for w in multi_match_warnings:
            print(f"  {w['name']} (excel rows: {w['rows']})")
            for c in w['candidates']:
                print(f"    候選 {c['chartNo']} 生日 {c['birthday']}")
        print()

    print('=== Orders by progress ===')
    by_progress = {}
    for o in orders:
        by_progress[o['progress']] = by_progress.get(o['progress'], 0) + 1
    for k, v in by_progress.items():
        print(f'  {k}: {v}')


if __name__ == '__main__':
    main()
