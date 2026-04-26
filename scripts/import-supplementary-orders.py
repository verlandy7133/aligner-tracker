# 補充下單匯入：從 補充下單紀錄.xlsx 抓「真正有副數區間的下單批次」
# 比對現有 patients-import.json + excel-patient-updates.json，
# 把同人重複偵測 + 跟主下單記錄合併。
#
# 兩階段：
#   1. dry-run 印出統計、重複名單、不確定的 case
#   2. 加 --apply 才真的寫進 dev-data/excel-orders.json + excel-patient-updates.json

import argparse
import json
import re
import sys
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import os
# 補充下單紀錄 Excel 路徑，可用環境變數 ALIGNER_SUPP_XLSX 覆蓋
EXCEL_PATH = os.environ.get('ALIGNER_SUPP_XLSX') or r'C:\Users\YOUR_USER\Downloads\補充下單紀錄.xlsx'
PATIENTS_JSON = Path('dev-data/patients-import.json')
PATIENT_UPDATES_JSON = Path('dev-data/excel-patient-updates.json')
ORDERS_JSON = Path('dev-data/excel-orders.json')

ALIGNER_PATTERN = re.compile(r'[UL]+\d+(?:-\d+)?(?:\+[UL]+\d+(?:-\d+)?)*')
PROGRESS_VALID = {'尚未開始', '已下單牙套', '診所已收到牙套', '已完成'}


def normalize_progress(v):
    if not v:
        return '尚未開始'
    s = str(v).strip()
    return s if s in PROGRESS_VALID else '尚未開始'


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


def normalize_name(raw):
    s = str(raw).strip()
    m = re.match(r'^(.+?)(\d{6,7})$', s)
    if not m:
        return s, None
    name = m.group(1).strip()
    digits = m.group(2)
    if len(digits) == 6:
        roc_y, mo, d = int(digits[:2]), int(digits[2:4]), int(digits[4:6])
    else:
        roc_y, mo, d = int(digits[:3]), int(digits[3:5]), int(digits[5:7])
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return s, None
    return name, f'{roc_y + 1911:04d}-{mo:02d}-{d:02d}'


def parse_aligner_range(text):
    if not text:
        return '', ''
    s = str(text).strip()
    m = ALIGNER_PATTERN.match(s)
    if not m:
        return '', s
    rng = m.group(0)
    rest = s[len(rng):].strip()
    return rng, rest


def extract_actual_order_date(text, fallback_year):
    if not text:
        return None
    m = re.search(r'(\d{1,2})/(\d{1,2})\s*下單', str(text))
    if not m:
        return None
    mo, d = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return f'{fallback_year}-{mo:02d}-{d:02d}'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--apply', action='store_true', help='實際寫入 (預設 dry-run)')
    args = ap.parse_args()

    # 載入現有資料
    patients_data = json.loads(PATIENTS_JSON.read_text(encoding='utf-8'))
    existing_patients = patients_data['patients']

    # 用 (name, birthday) 當 key
    name_birthday_to_pat = {}
    name_to_pats = defaultdict(list)
    for p in existing_patients:
        name_birthday_to_pat[(p['name'], p['birthday'])] = p
        name_to_pats[p['name']].append(p)

    # 載入既有 orders / updates (要避免重複塞)
    existing_orders = []
    if ORDERS_JSON.exists():
        existing_orders = json.loads(ORDERS_JSON.read_text(encoding='utf-8'))['orders']
    existing_order_keys = set(
        (o['patientId'], o['date'], o['alignerRange']) for o in existing_orders
    )

    existing_updates = []
    if PATIENT_UPDATES_JSON.exists():
        ud = json.loads(PATIENT_UPDATES_JSON.read_text(encoding='utf-8'))
        existing_updates = ud.get('updates', [])
    existing_new_patients = []
    if PATIENT_UPDATES_JSON.exists():
        ud = json.loads(PATIENT_UPDATES_JSON.read_text(encoding='utf-8'))
        existing_new_patients = ud.get('newPatients', [])
    existing_new_pat_names = set(p['name'] for p in existing_new_patients)

    # 讀補充 Excel
    wb = load_workbook(EXCEL_PATH, data_only=True)
    sh = wb.active

    # 收集只有 UL/LL 區間的 row
    candidate_rows = []
    for r in range(2, sh.max_row + 1):
        raw_name = sh.cell(row=r, column=2).value
        if not raw_name:
            continue
        notes_col = sh.cell(row=r, column=8).value or ''
        rng, rest = parse_aligner_range(notes_col)
        if not rng:
            continue  # 沒副數區間 → 不算下單
        candidate_rows.append({
            'r': r,
            'rawName': str(raw_name).strip(),
            'date': parse_date(sh.cell(row=r, column=1).value),
            'doctor': sh.cell(row=r, column=3).value,
            'progress': sh.cell(row=r, column=9).value,
            'expectedDate': sh.cell(row=r, column=15).value,
            'actualDate': sh.cell(row=r, column=17).value,
            'nextStep': sh.cell(row=r, column=18).value,
            'notes': sh.cell(row=r, column=10).value,
            'alignerRange': rng,
            'inlineNotes': rest,
        })

    # 另外掃「所有有姓名的 row」用來補醫師（不限有副數區間）
    # 因為很多 row 是回診/收件紀錄，沒副數區間但仍記錄醫師
    all_doctor_rows = []
    for r in range(2, sh.max_row + 1):
        raw_name = sh.cell(row=r, column=2).value
        doctor = sh.cell(row=r, column=3).value
        if not raw_name or not doctor:
            continue
        all_doctor_rows.append({
            'r': r,
            'rawName': str(raw_name).strip(),
            'doctor': str(doctor).strip(),
        })

    print(f'=== 補充下單紀錄 dry-run (apply={args.apply}) ===')
    print(f'Excel 內有 UL/LL 區間的 row：{len(candidate_rows)}')
    print(f'Excel 內有姓名+醫師的 row：{len(all_doctor_rows)}')

    # 同檔內重複名偵測
    by_name_within_supp = defaultdict(list)
    for row in candidate_rows:
        clean_name, _ = normalize_name(row['rawName'])
        by_name_within_supp[clean_name].append(row)

    intra_dups = {n: rows for n, rows in by_name_within_supp.items() if len(rows) > 1}
    print(f'同檔內出現 >1 次的姓名：{len(intra_dups)} 個 (涉及 {sum(len(r) for r in intra_dups.values())} 筆)')
    if intra_dups:
        print('  詳細：')
        for n, rows in list(intra_dups.items())[:15]:
            print(f"    {n[0]+'*'*(len(n)-1)}: {len(rows)} 筆 → rows {[r['r'] for r in rows]}, ranges {[r['alignerRange'] for r in rows]}")

    new_orders = []
    patient_updates_to_apply = []  # 套到既有 patient
    new_patients_to_create = []
    skipped_dup = 0
    skipped_no_match_no_birthday = []

    # 收集既有 patient 的醫師回填資訊（每位病人取第一個有醫師值的 row）
    # 用 (existing patient id) -> {'doctor': '...'} 暫存
    supp_doctor_by_pid = {}

    case_year = 2026

    for row in candidate_rows:
        clean_name, hint_birthday = normalize_name(row['rawName'])

        # 找 patient
        patient = None
        if hint_birthday:
            patient = name_birthday_to_pat.get((clean_name, hint_birthday))
        if not patient:
            candidates = name_to_pats.get(clean_name, [])
            if len(candidates) == 1:
                patient = candidates[0]
            elif len(candidates) > 1:
                # 多人同名，沒有 birthday hint → 跳過讓 user 處理
                skipped_no_match_no_birthday.append({
                    'name': clean_name,
                    'row': row['r'],
                    'candidates': [c['chartNo'] for c in candidates],
                })
                continue

        if not patient:
            # 新病患 — 看是否已建過
            if clean_name in existing_new_pat_names:
                # 已經是 excel-patient-updates 裡新建的人
                patient = next(p for p in existing_new_patients if p['name'] == clean_name)
            else:
                # 真新人，先建 patient (用既有 max chartNo + 自增)
                max_chart = max(int(p['chartNo']) for p in existing_patients + new_patients_to_create + existing_new_patients)
                new_pat = {
                    'id': str(uuid.uuid4()),
                    'chartNo': f'{max_chart + 1:04d}',
                    'name': clean_name,
                    'birthday': hint_birthday,
                    'productLine': 'riyue',
                    'status': 'active',
                    'orderDate': row['date'],
                    'startDate': None,
                    'totalAlignersUpper': None,
                    'currentAlignerUpper': None,
                    'totalAlignersLower': None,
                    'currentAlignerLower': None,
                    'cycleDays': 10,
                    'lastVisit': None,
                    'nextVisit': None,
                    'hasConsent': False,
                    'consentPdfPath': None,
                    'scanInfo': None,
                    'doctor': str(row['doctor']).strip() if row['doctor'] else None,
                    'flags': [],
                    'notes': '(從 補充下單紀錄.xlsx 匯入新建)',
                    'sourceFolder': '',
                    'createdAt': datetime.now().isoformat(),
                    'updatedAt': datetime.now().isoformat(),
                }
                new_patients_to_create.append(new_pat)
                patient = new_pat
                existing_new_pat_names.add(clean_name)

        # 構造 order
        actual_date = extract_actual_order_date(
            f"{row['inlineNotes'] or ''} {row['notes'] or ''}", case_year
        )
        order_date = actual_date or row['date'] or ''

        order_key = (patient['id'], order_date, row['alignerRange'])
        if order_key in existing_order_keys:
            skipped_dup += 1
            continue
        existing_order_keys.add(order_key)

        new_orders.append({
            'id': f"supp-{row['r']:04d}-{patient['id']}",
            'patientId': patient['id'],
            'patientChartNo': patient['chartNo'],
            'patientName': patient['name'],
            'date': order_date,
            'doctor': str(row['doctor']).strip() if row['doctor'] else (patient.get('doctor') or ''),
            'batchType': '',
            'alignerRange': row['alignerRange'],
            'progress': normalize_progress(row['progress']),
            'expectedDate': parse_date(row['expectedDate']),
            'actualDate': parse_date(row['actualDate']),
            'nextStep': str(row['nextStep']).strip() if row['nextStep'] else '',
            'notes': ' '.join(filter(None, [
                row['inlineNotes'] or '',
                str(row['notes']).strip() if row['notes'] else '',
                '(補充)',
            ])).strip(),
            'lab': '美鉑',  # 預設，user 之後可改
            'createdAt': datetime.now().isoformat(),
            'updatedAt': datetime.now().isoformat(),
        })

    # 從 all_doctor_rows match 既有 patient，收集醫師資訊
    # 同名多人 + 沒生日 hint → 跳過 (跟 candidate_rows 一樣的 match 邏輯)
    skipped_doctor_no_match = 0
    for row in all_doctor_rows:
        clean_name, hint_birthday = normalize_name(row['rawName'])
        patient = None
        if hint_birthday:
            patient = name_birthday_to_pat.get((clean_name, hint_birthday))
        if not patient:
            cands = name_to_pats.get(clean_name, [])
            if len(cands) == 1:
                patient = cands[0]
            else:
                skipped_doctor_no_match += 1
                continue
        # 第一筆即取，後續同 patient 的 row 不覆寫
        if patient['id'] not in supp_doctor_by_pid:
            supp_doctor_by_pid[patient['id']] = row['doctor']

    # 合併到既有 updates：補充檔的醫師資訊只填空缺，不覆寫主檔已有的醫師
    existing_upd_by_id = {u['id']: u for u in existing_updates}
    pat_lookup = {p['id']: p for p in existing_patients}
    new_doctor_fills = 0
    skipped_existing_doctor = 0
    for pid, dr in supp_doctor_by_pid.items():
        cur = existing_upd_by_id.get(pid)
        if cur and cur.get('doctor'):
            skipped_existing_doctor += 1
            continue
        ref = pat_lookup.get(pid, {})
        if cur:
            cur['doctor'] = dr
            cur.setdefault('refName', ref.get('name'))
            cur.setdefault('refBirthday', ref.get('birthday'))
        else:
            existing_upd_by_id[pid] = {
                'id': pid,
                'refName': ref.get('name'),
                'refBirthday': ref.get('birthday'),
                'doctor': dr,
            }
        new_doctor_fills += 1

    print()
    print('=== 結果 ===')
    print(f'  將新增 orders: {len(new_orders)}')
    print(f'  將新建 patients: {len(new_patients_to_create)}')
    print(f'  與現有 orders 重複跳過: {skipped_dup}')
    print(f'  同名多人需手動確認 (跳過): {len(skipped_no_match_no_birthday)}')
    print(f'  新增/補上醫師到既有 patient: {new_doctor_fills}')
    print(f'  既有已有醫師、跳過: {skipped_existing_doctor}')
    if skipped_no_match_no_birthday:
        for s in skipped_no_match_no_birthday[:10]:
            print(f"    {s['name']} (row {s['row']}) 候選 chartNo: {s['candidates']}")

    if new_patients_to_create:
        print()
        print('=== 自動補建的新病患 (前 20) ===')
        for p in new_patients_to_create[:20]:
            print(f"  {p['chartNo']} {p['name']} (生日={p['birthday']}, 醫師={p['doctor']})")

    if not args.apply:
        print()
        print('### dry-run 結束。要實際寫入請加 --apply ###')
        return

    # ── 寫入 ──
    # 1. orders
    if ORDERS_JSON.exists():
        existing = json.loads(ORDERS_JSON.read_text(encoding='utf-8'))
        existing['orders'].extend(new_orders)
        existing['count'] = len(existing['orders'])
        ORDERS_JSON.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding='utf-8')
    # 2. new patients + 醫師回填 (擴充 excel-patient-updates)
    if PATIENT_UPDATES_JSON.exists():
        upd = json.loads(PATIENT_UPDATES_JSON.read_text(encoding='utf-8'))
        upd.setdefault('newPatients', []).extend(new_patients_to_create)
        upd['newPatientsCount'] = len(upd['newPatients'])
        # 用合併後的 updates 取代 (existing_upd_by_id 已包含原本 + 新補的醫師)
        upd['updates'] = list(existing_upd_by_id.values())
        upd['count'] = len(upd['updates'])
        PATIENT_UPDATES_JSON.write_text(json.dumps(upd, ensure_ascii=False, indent=2), encoding='utf-8')

    print()
    print(f'✓ 寫入完成。重整 App 並到 /設定 → 清空 DB 重 seed 即可看到新資料。')


if __name__ == '__main__':
    main()
